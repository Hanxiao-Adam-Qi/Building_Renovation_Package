from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\PV_Data.xlsx', data_only=True)
wb2 = Workbook()
sheet2 = wb2.create_sheet('electricity savings', index=0)
sheet3 = wb2.create_sheet('electricity saving annually')

for i in range(1, 13):
    sheet = wb['Blad{}'.format(i)]
    max_row = sheet.max_row
    if max_row>749:
        max_row = 749
    print(max_row)

    for l in range(6, max_row+1):
        for j in ['AE', 'AG']:
            max_row3 = sheet3.max_row
            if j == 'AE':
                max_row3 = sheet3.max_row
                prod_small = sheet['{}{}'.format(j,l)].value
                sheet2.cell(l-3, 2*i-1).value = str(prod_small)
                sheet3.cell(max_row3+1, 1).value = str(prod_small)
            else:
                prod_large = sheet['{}{}'.format(j,l)].value
                sheet2.cell(l - 3, 2 * i).value = str(prod_large)
                sheet3.cell(max_row3, 2).value = str(prod_large)

wb2.save(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\PV_Data2.0.xlsx')