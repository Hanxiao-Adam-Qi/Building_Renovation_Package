import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
starttime = datetime.datetime.now()


month_hour = [744, 672, 744, 720, 744, 720, 744, 744, 720, 744, 720, 744]

wb_simulation_use = load_workbook(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\elec_use_hourly.xlsx',read_only = True)
wb_initial_use = load_workbook(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\LCC.xlsx',read_only = True)
sheet_simulation_use = wb_simulation_use['electricity use hourly']

wb_PV_price = load_workbook(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\PV_prices.xlsx', read_only = True, data_only = True)
sheet_PV_production = wb_PV_price['Hourly profiles']


# get prices for months
selling_price = []
buying_price = []
tax_compen = []
for ii in range(1, 13):
    sheet_price = wb_PV_price['Blad{}'.format(ii)]
    selling_price.append(sheet_price['AP7'].value)
    buying_price.append(sheet_price['AP14'].value)
    tax_compen.append(sheet_price['AP16'].value)


# wb_result = Workbook()
# sheet_initial = wb_result.create_sheet('initial use', index=0)

# get annually production
prod_lar = []
prod_smal = []
for zz in range(sheet_PV_production.max_row-2):
    prod_lar1 = sheet_PV_production.cell(zz+3, 2).value
    prod_smal1 = sheet_PV_production.cell(zz+3, 4).value
    prod_lar.append(prod_lar1)
    prod_smal.append(prod_smal1)

# for 140 simulation
for mm in range(sheet_simulation_use.max_column-1):
    simucase = []
    for dd in range(sheet_simulation_use.max_row-1):
        simucase_1 = sheet_simulation_use.cell(dd+2, mm+2).value    # start from basecase simulation
        simucase.append(simucase_1)                                 # Should we use bills or simulations?


    # separate annual data into monthly
    monthly_savings_kWh = []
    monthly_savings_SEK = []
    m = 0
    for i in month_hour:
        m = m+i
        z = m-i

        # separate simulation data monthly
        monthly_use_simulation = simucase[z:m]
        print(monthly_use_simulation)

        # separate produce data monthly
        prod_lar_monthly = prod_lar[z:m]
        prod_smal_monthly = prod_smal[z:m]

        # Large system---calculate selling and buying electricity amount
        lar_sold_kWh = 0
        lar_buy_kWh = 0
        for xx in range(len(prod_lar_monthly)):
            if prod_lar_monthly[xx] - monthly_use_simulation[xx] > 0:
                lar_sold_kWh = lar_sold_kWh + (prod_lar_monthly[xx] - monthly_use_simulation[xx])
            else:
                lar_buy_kWh = lar_buy_kWh + (prod_lar_monthly[xx] - monthly_use_simulation[xx])

    print(sheet_simulation_use.cell(1, mm+3).value)
    print(lar_sold_kWh, lar_buy_kWh)



endtime = datetime.datetime.now()
print (endtime - starttime)

# wb_result.save(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\Results.xlsx')