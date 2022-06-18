import xlrd
import xlwt
import xlutils.copy
from openpyxl import Workbook
from openpyxl import load_workbook

# # xlwt
# wb_el_use0 = xlwt.Workbook()
# wb_el_use0.add_sheet('elec use',cell_overwrite_ok = True)
# wb_el_use0.save(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\elec_use_hourly.xls')

Insulation_Type = ['Glasswool', 'SIP']
Insulation_Thinkness = [0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3]
Window_Types = [0,1, 2, 3, 4, 5, 6]
HP_COP = [1, 3]

years = 30
DH_price = 0.854
DH_g = 0.01
DH_i = 0.03
DH_ini_fix = 4542.108

wb_elec_use1 = Workbook()
sheet1 = wb_elec_use1.create_sheet('electricity use hourly', index=0)
sheet1.cell(1, 1, 'electricity use hourly')

for i in Insulation_Type:
    if i == Insulation_Type[0]:
        for j in Insulation_Thinkness:
            for k in Window_Types:
                for l in HP_COP:
                    wb1 = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\python_grasshopper data_elec use\{}-{}_Window-{}_HCOP-{}.xls'.format(i, j, k, l))
                    obj1 = wb1.sheet_by_name('Simulation electricity use')

                    col1 = sheet1.max_column
                    # wb_el_use = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\elec_use_hourly.xls')
                    # sheet1 = wb_el_use.sheet_by_name('elec use')
                    for rows in range(obj1.nrows):
                        a = obj1.cell_value(rows, 0)
                        sheet1.cell(rows+1, col1+1, a)
                    print('{}-{}_Window-{}_HCOP-{} is down'.format(i, j, k, l))
                    print(col1)


    else:
        for j in Insulation_Thinkness[1:4]:
            for k in Window_Types:
                for l in HP_COP:
                    wb2 = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\python_grasshopper data_elec use\{}-{}_Window-{}_HCOP-{}.xls'.format(i, j, k, l))
                    obj2 = wb2.sheet_by_name('Simulation electricity use')

                    col2 = sheet1.max_column
                    for rows in range(obj2.nrows):
                        a = obj2.cell_value(rows, 0)
                        sheet1.cell(rows + 1, col2 + 1, a)
                    print('{}-{}_Window-{}_HCOP-{} is down'.format(i, j, k, l))
                    print(col2)

wb_elec_use1.save(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\elec_use_hourly.xlsx')


    #                 b = []
    #                 wb2 = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\Energy_simu\python_grasshopper data_elec use\{}-{}_Window-{}_HCOP-{}.xls'.format(i, j, k, l))
    #                 obj2 = wb2.sheet_by_index(0)
    #                 for rows in range(obj2.nrows):
    #                     b.append(obj2.col_values(0, 0, obj2.nrows))
    #                 wb_el_use.create_sheet('{}-{}_Window-{}_HCOP-{}'.format(i, j, k, l))
    #                 active_ws.cell(1, u + 1).value = str(b[u])
    #                 # eu.append(b)
    #                 print('{}-{}_Window-{}_HCOP-{} is down'.format(i, j, k, l))

# for u in range(len(eu)):
#     for v in range(len(eu[u])):
#         hourly_elec_use = eu[u][v]
#         sheet.cell(v, u).value = str(hourly_elec_use)
#         print('writing row:{}/8761;  column:{}/140')




