import xlrd
from openpyxl import Workbook

Insulation_Type = ['Glasswool', 'SIP']
Insulation_Thinkness = [0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3]
Window_Types = [0,1, 2, 3, 4, 5, 6]
HP_COP = [1]

# # test
# Insulation_Type = ['Glasswool']
# Insulation_Thinkness = [0]
# Window_Types = [0]
# HP_COP = [1]

wb_heat_use1 = Workbook()
sheet1 = wb_heat_use1.create_sheet('Heating use hourly', index=0)
sheet1.cell(1, 1, 'Heating use hourly')

for i in Insulation_Type:
    if i == Insulation_Type[0]:
        for j in Insulation_Thinkness:
            for k in Window_Types:
                for l in HP_COP:
                    wb1 = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\Heating data\{}-{}_Window-{}_HCOP-{}.xls'.format(i, j, k, l))
                    obj1 = wb1.sheet_by_name('{}-{}_Window-{}_HCOP-{}'.format(i, j, k, l))

                    col1 = sheet1.max_column

                    for rows in range(obj1.nrows):
                        heat_use = obj1.cell_value(rows, 0)
                        sheet1.cell(rows+1, col1+1, heat_use)
                    print('{}-{}_Window-{}_HCOP-{} is down'.format(i, j, k, l))
                    print(col1)

    else:
        for j in Insulation_Thinkness[1:4]:
            for k in Window_Types:
                for l in HP_COP:
                    wb2 = xlrd.open_workbook(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\Heating data\{}-{}_Window-{}_HCOP-{}.xls'.format(i, j, k, l))
                    obj2 = wb2.sheet_by_name('{}-{}_Window-{}_HCOP-{}'.format(i, j, k, l))
#
                    col1 = sheet1.max_column
                    for rows in range(obj2.nrows):
                        heat_use = obj2.cell_value(rows, 0)
                        sheet1.cell(rows + 1, col1 + 1, heat_use)
                    print('{}-{}_Window-{}_HCOP-{} is down'.format(i, j, k, l))
                    print(col1)

wb_heat_use1.save(r'C:\Users\Adam\Desktop\AEBN20\LCC Calc\Heating hourly use.xlsx')
#
# # Unit is J. if /3600000 to kWh, number is wrong
